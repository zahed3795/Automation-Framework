import os
import json
import openpyxl


def json_data_reader(attribute):
    JSON_FILE = os.path.dirname(os.path.realpath('environment.json'.__file__))
    with open(JSON_FILE) as json_path:
        json_file = json.load(json_path)
    return json_file[attribute]

    # print(json_data_reader("browser"))


def xlsx_data_reader(rownum, columnno):
    sheetNM = 'credentials'
    XLSX_FILE = os.path.dirname(os.path.realpath('environment.xlsx'.__file__))
    workbook = openpyxl.load_workbook(XLSX_FILE)
    workbook_value = workbook[sheetNM].cell(row=rownum, column=columnno).value
    return workbook_value


def xlsx_data__writer(data, rownum, columnno):
    XLSX_FILE = os.path.dirname(os.path.realpath('environment.xlsx'.__file__))
    workbook = openpyxl.load_workbook(XLSX_FILE)
    sheetNM = 'credentials'
    workbook_value = workbook[sheetNM].cell(row=rownum, column=columnno).value = data
    return workbook.save(XLSX_FILE)




# print(data_reader.xlsx_data_reader(4,2))
